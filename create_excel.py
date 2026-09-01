import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook & worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "WBS & Rencana Kerja"
ws.views.sheetView[0].showGridLines = True

# Soft Green Color Palette
DARK_GREEN_FILL = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid") # Main Header
HEADER_GREEN_FILL = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid") # Table Header
SOFT_GREEN_LIGHT = PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid") # Alternating Rows
SOFT_GREEN_CARD = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid") # Stat Card Fill
STATUS_DONE_FILL = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid") # Soft Done

# Fonts
FONT_TITLE = Font(name="Segoe UI", size=15, bold=True, color="FFFFFF")
FONT_SUBTITLE = Font(name="Segoe UI", size=10, italic=True, color="E8F5E9")
FONT_TH = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
FONT_BOLD = Font(name="Segoe UI", size=10, bold=True, color="1B5E20")
FONT_REGULAR = Font(name="Segoe UI", size=10, color="212121")
FONT_CARD_TITLE = Font(name="Segoe UI", size=9, bold=True, color="388E3C")
FONT_CARD_VAL = Font(name="Segoe UI", size=14, bold=True, color="1B5E20")
FONT_DONE = Font(name="Segoe UI", size=10, bold=True, color="2E7D32")

# Alignments
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

# Borders
THIN_BORDER_COLOR = "C8E6C9"
thin_side = Side(border_style="thin", color=THIN_BORDER_COLOR)
BORDER_CELL = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
BORDER_CARD = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

# --- 1. TITLE BANNER ---
ws.merge_cells("A1:H2")
title_cell = ws["A1"]
title_cell.value = "RENCANA KERJA & WBS PEMBUATAN WEBSITE SALON YALIA BEAUTY"
title_cell.font = FONT_TITLE
title_cell.fill = DARK_GREEN_FILL
title_cell.alignment = ALIGN_CENTER

ws.merge_cells("A3:H3")
sub_cell = ws["A3"]
sub_cell.value = "Sistem Informasi Reservasi Salon & Home Service (Laravel 13 & Tailwind CSS) • Soft Green Theme"
sub_cell.font = FONT_SUBTITLE
sub_cell.fill = DARK_GREEN_FILL
sub_cell.alignment = ALIGN_CENTER

# --- 2. SUMMARY DASHBOARD CARDS ---
cards_config = [
    ("A5:B5", "A6:B6", "TOTAL TUGAS (WBS)", '=COUNTA(B10:B33)'),
    ("C5:D5", "C6:D6", "SELESAI (100%)", '=COUNTIF(G10:G33, "Selesai")'),
    ("E5:F5", "E6:F6", "DALAM PROSES", '=COUNTIF(G10:G33, "Dalam Proses")'),
    ("G5:H5", "G6:H6", "TOTAL PROGRESS", '=AVERAGE(H10:H33)')
]

for top_range, val_range, lbl, formula in cards_config:
    ws.merge_cells(top_range)
    ws.merge_cells(val_range)
    
    top_cell = ws[top_range.split(":")[0]]
    top_cell.value = lbl
    top_cell.font = FONT_CARD_TITLE
    top_cell.fill = SOFT_GREEN_CARD
    top_cell.alignment = ALIGN_CENTER
    
    val_cell = ws[val_range.split(":")[0]]
    val_cell.value = formula
    val_cell.font = FONT_CARD_VAL
    val_cell.fill = SOFT_GREEN_CARD
    val_cell.alignment = ALIGN_CENTER
    if "PROGRESS" in lbl:
        val_cell.number_format = '0.0%'

    # Border around card cells
    top_start_col, top_start_row = openpyxl.utils.coordinate_to_tuple(top_range.split(":")[0])
    val_end_col, val_end_row = openpyxl.utils.coordinate_to_tuple(val_range.split(":")[1])
    for r in range(top_start_row, val_end_row + 1):
        for c in range(top_start_col, val_end_col + 1):
            ws.cell(row=r, column=c).border = BORDER_CARD
            ws.cell(row=r, column=c).fill = SOFT_GREEN_CARD

# --- 3. TABLE HEADERS ---
headers = [
    "No", "Kode WBS", "Nama Tugas / Modul Sistem", "Fase Project",
    "PIC / Role", "Durasi", "Status", "Progress (%)"
]

header_row = 9
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=header_row, column=col_num)
    cell.value = header
    cell.font = FONT_TH
    cell.fill = HEADER_GREEN_FILL
    cell.alignment = ALIGN_CENTER
    cell.border = BORDER_CELL

ws.row_dimensions[header_row].height = 28

# --- 4. TASK DATA (24 TASKS) ---
tasks = [
    (1, "WBS 1.1", "Analisis Kebutuhan Sistem & User Persona", "Fase 1: Perancangan", "System Analyst", 2, "Selesai", 1.0),
    (2, "WBS 1.2", "Pembuatan Flowchart Manual Booking & Gamifikasi", "Fase 1: Perancangan", "System Analyst", 3, "Selesai", 1.0),
    (3, "WBS 1.3", "Perancangan ERD Pohon Diagrams (15 Entitas)", "Fase 1: Perancangan", "Database Engineer", 3, "Selesai", 1.0),
    (4, "WBS 1.4", "Perancangan DFD Level 0 & DFD Level 1 (Data Stores)", "Fase 1: Perancangan", "System Analyst", 2, "Selesai", 1.0),
    
    (5, "WBS 2.1", "Penyusunan Design System & Palette (Soft Pink #f472b6)", "Fase 2: UI/UX", "UI/UX Designer", 2, "Selesai", 1.0),
    (6, "WBS 2.2", "Layouting Landing Page & Catalog Treatment", "Fase 2: UI/UX", "Frontend Dev", 4, "Selesai", 1.0),
    (7, "WBS 2.3", "Sistem Komponen Navigation & Responsive Header/Footer", "Fase 2: UI/UX", "Frontend Dev", 3, "Selesai", 1.0),
    (8, "WBS 2.4", "Halaman Admin Dashboard & Panel Manajemen", "Fase 2: UI/UX", "Frontend Dev", 4, "Selesai", 1.0),
    
    (9, "WBS 3.1", "Setup Laravel 13 Framework & Migrasi Database", "Fase 3: Backend", "Backend Dev", 2, "Selesai", 1.0),
    (10, "WBS 3.2", "Implementasi Eloquent Models & Relasi Database", "Fase 3: Backend", "Backend Dev", 3, "Selesai", 1.0),
    (11, "WBS 3.3", "Fitur Autentikasi Breeze & Google OAuth Integration", "Fase 3: Backend", "Backend Dev", 3, "Selesai", 1.0),
    (12, "WBS 3.4", "Manajemen Treatment, Kategori, & Beautician Profile", "Fase 3: Backend", "Backend Dev", 3, "Selesai", 1.0),
    
    (13, "WBS 4.1", "Algoritma Perhitungan Slot Jam Availability", "Fase 4: Core Engine", "Backend Dev", 4, "Selesai", 1.0),
    (14, "WBS 4.2", "Fitur Reservasi Salon vs Home Service Radius GPS", "Fase 4: Core Engine", "Fullstack Dev", 4, "Selesai", 1.0),
    (15, "WBS 4.3", "Modul Penugasan Kapster & Beautician Schedule", "Fase 4: Core Engine", "Backend Dev", 3, "Selesai", 1.0),
    (16, "WBS 4.4", "Modul Batal Booking & Reschedule Jadwal Perawatan", "Fase 4: Core Engine", "Fullstack Dev", 3, "Selesai", 1.0),
    
    (17, "WBS 5.1", "Integrasi Midtrans Payment Gateway (QRIS & Snap)", "Fase 5: Integration", "Backend Dev", 4, "Selesai", 1.0),
    (18, "WBS 5.2", "Webhook Handler Auto Settlement & Status Pembayaran", "Fase 5: Integration", "Backend Dev", 2, "Selesai", 1.0),
    (19, "WBS 5.3", "Sistem Gamifikasi Poin Reward PTS & Tier Loyalty", "Fase 5: Integration", "Fullstack Dev", 4, "Selesai", 1.0),
    (20, "WBS 5.4", "Modul Kupon Voucher Diskon & Check-in Streak", "Fase 5: Integration", "Fullstack Dev", 3, "Selesai", 1.0),
    
    (21, "WBS 6.1", "Unit Testing & Integration Test Suite (PHPUnit)", "Fase 6: Testing", "QA Engineer", 3, "Selesai", 1.0),
    (22, "WBS 6.2", "Optimalisasi Performance Performa Laravel Octane v2", "Fase 6: Testing", "DevOps Engineer", 2, "Selesai", 1.0),
    (23, "WBS 6.3", "Keamanan Data, Middleware, & Input Sanitization", "Fase 6: Testing", "Security Lead", 2, "Selesai", 1.0),
    (24, "WBS 6.4", "Final Deployment Staging Server & UAT User Review", "Fase 6: Testing", "Project Manager", 2, "Selesai", 1.0),
]

start_row = 10
for i, task in enumerate(tasks):
    current_row = start_row + i
    ws.row_dimensions[current_row].height = 22
    
    no, wbs, name, phase, pic, dur, status, prog = task
    
    row_data = [
        (no, ALIGN_CENTER, FONT_REGULAR),
        (wbs, ALIGN_CENTER, FONT_BOLD),
        (name, ALIGN_LEFT, FONT_REGULAR),
        (phase, ALIGN_LEFT, FONT_REGULAR),
        (pic, ALIGN_LEFT, FONT_REGULAR),
        (f"{dur} Hari", ALIGN_CENTER, FONT_REGULAR),
        (status, ALIGN_CENTER, FONT_DONE),
        (prog, ALIGN_CENTER, FONT_DONE)
    ]
    
    # Alternating fill
    row_fill = SOFT_GREEN_LIGHT if i % 2 == 1 else None
    
    for col_num, (val, align, font) in enumerate(row_data, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = val
        cell.alignment = align
        cell.font = font
        cell.border = BORDER_CELL
        
        if col_num == 7: # Status Column
            cell.fill = STATUS_DONE_FILL
        elif col_num == 8: # Progress Column
            cell.number_format = '0.0%'
            cell.fill = STATUS_DONE_FILL
        elif row_fill:
            cell.fill = row_fill

# --- 5. AUTO-FIT COLUMN WIDTHS ---
column_widths = {
    'A': 8,   # No
    'B': 14,  # Kode WBS
    'C': 54,  # Nama Tugas / Modul
    'D': 24,  # Fase Project
    'E': 22,  # PIC / Role
    'F': 14,  # Durasi
    'G': 16,  # Status
    'H': 16   # Progress (%)
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# --- 6. SAVE WORKBOOK ---
output_path = r"d:\laragon\www\salon-yalia-beauty\Rencana_Kerja_Salon_Yalia_Beauty.xlsx"
wb.save(output_path)
print(f"Excel generated successfully at: {output_path}")
