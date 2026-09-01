import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Create workbook & worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Estimasi Biaya SMK Freelance"
ws.views.sheetView[0].showGridLines = True

# Color Palette: Soft Sage & Slate (Minimalist Corporate)
COLOR_HEADER_BG = PatternFill(start_color="2D5A27", end_color="2D5A27", fill_type="solid")  # Muted Dark Sage Green
COLOR_TABLE_BG = PatternFill(start_color="385723", end_color="385723", fill_type="solid")   # Table Header Sage
COLOR_ZEBRA_BG = PatternFill(start_color="F4F9F3", end_color="F4F9F3", fill_type="solid")   # Soft Tint
COLOR_TOTAL_BG = PatternFill(start_color="E2EFE0", end_color="E2EFE0", fill_type="solid")   # Accent Soft Green

# Fonts
FONT_COMPANY = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
FONT_SUBTITLE = Font(name="Segoe UI", size=10, italic=True, color="E2EFE0")
FONT_TH = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
FONT_SECTION = Font(name="Segoe UI", size=10, bold=True, color="2D5A27")
FONT_REGULAR = Font(name="Segoe UI", size=10, color="212121")
FONT_BOLD = Font(name="Segoe UI", size=10, bold=True, color="212121")
FONT_TOTAL = Font(name="Segoe UI", size=11, bold=True, color="1B4314")
FONT_LABEL = Font(name="Segoe UI", size=9, bold=True, color="555555")

# Alignments
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

# Borders
thin_gray = Side(border_style="thin", color="D9D9D9")
double_bottom = Side(border_style="double", color="2D5A27")

BORDER_CELL = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
BORDER_TOTAL = Border(top=thin_gray, bottom=double_bottom, left=thin_gray, right=thin_gray)

# Currency Format
CURRENCY_FORMAT = 'Rp #,##0'

# --- 1. HEADER BANNER ---
ws.merge_cells("A1:G2")
t_cell = ws["A1"]
t_cell.value = "PENWARAN HARGA & ESTIMASI BIAYA DEVELOPMENT WEBSITE (FRESH GRADUATE SMK RATE)"
t_cell.font = FONT_COMPANY
t_cell.fill = COLOR_HEADER_BG
t_cell.alignment = ALIGN_CENTER

ws.merge_cells("A3:G3")
sub_cell = ws["A3"]
sub_cell.value = "Proyek: Sistem Informasi Reservasi Salon & Home Service (Salon Yalia Beauty) • Budget SMK 5-7 Juta"
sub_cell.font = FONT_SUBTITLE
sub_cell.fill = COLOR_HEADER_BG
sub_cell.alignment = ALIGN_CENTER

# --- 2. INFORMASI PROYEK & KLIEN ---
info_data = [
    ("Klien", "Salon Yalia Beauty", "No. Dokumen", "QUO-2026/YB/SMK-01"),
    ("Perihal", "Pengembangan Website & Sistem Reservasi", "Tanggal", "31 Agustus 2026"),
    ("Developer / Role", "Junior Fullstack Dev (Fresh Grad SMK RPL)", "Masa Berlaku", "30 Hari")
]

for idx, (lbl1, val1, lbl2, val2) in enumerate(info_data, 5):
    ws.cell(row=idx, column=1, value=lbl1).font = FONT_LABEL
    ws.cell(row=idx, column=1).alignment = ALIGN_LEFT
    ws.cell(row=idx, column=2, value=val1).font = FONT_BOLD
    ws.cell(row=idx, column=2).alignment = ALIGN_LEFT
    
    ws.cell(row=idx, column=5, value=lbl2).font = FONT_LABEL
    ws.cell(row=idx, column=5).alignment = ALIGN_LEFT
    ws.cell(row=idx, column=6, value=val2).font = FONT_REGULAR
    ws.cell(row=idx, column=6).alignment = ALIGN_LEFT

ws.row_dimensions[8].height = 10

# --- 3. TABEL RINCIAN BIAYA FITUR ---
headers = ["No", "Modul / Fitur Website", "Deskripsi & Cakupan Spesifikasi", "Qty", "Satuan", "Harga Satuan (Rp)", "Total Biaya (Rp)"]
header_row = 9
ws.row_dimensions[header_row].height = 26

for col_num, h in enumerate(headers, 1):
    cell = ws.cell(row=header_row, column=col_num)
    cell.value = h
    cell.font = FONT_TH
    cell.fill = COLOR_TABLE_BG
    cell.alignment = ALIGN_CENTER
    cell.border = BORDER_CELL

# Module / Feature Items (Target: 6.000.000 IDR)
items = [
    (1, "Dokumentasi & Arsitektur Sistem", 
     "Perancangan Flowchart Manual (5 Diagram), ERD Pohon (15 Entitas), DFD Level 0 & Level 1 Data Stores, Dokumen UKK RPL",
     1, "Paket", 750000),
    
    (2, "Design UI/UX & Responsive Front-End", 
     "Layouting Landing Page Soft Pink, Katalog Treatment, Navigation Component, Responsive Mobile & Desktop (Blade & Tailwind)",
     1, "Paket", 1000000),
    
    (3, "Core Booking Engine & Home Service GPS", 
     "Sistem Reservasi Salon vs Home Service, Calculation Slot Jam Bebas Bentrok, Validasi Radius GPS, Penugasan Beautician",
     1, "Modul", 1500000),
    
    (4, "Payment Gateway & Webhook Midtrans", 
     "Integrasi Payment Gateway Midtrans (QRIS & Snap Popup), Webhook Listener Settlement/Expired, Digital E-Receipt",
     1, "Modul", 1000000),
    
    (5, "Gamifikasi, Loyalty Tier & Voucher Engine", 
     "Sistem Poin Rewards PTS per Treatment, Badge Level (Silver/Gold/Platinum), Check-in Streak, Kode Voucher Diskon",
     1, "Modul", 750000),
    
    (6, "Admin Portal, Journal Finance & OCR Receipt", 
     "Dashboard Admin Monitoring Reservasi, Laporan Keuangan Ledger (Income/Expense), OCR AI Ekstraksi Struk Belanja",
     1, "Modul", 1000000),
]

start_row = 10
for i, item in enumerate(items):
    curr_row = start_row + i
    ws.row_dimensions[curr_row].height = 32
    no, name, desc, qty, unit, price = item
    
    row_fill = COLOR_ZEBRA_BG if i % 2 == 1 else None
    
    # Cells assignment
    c1 = ws.cell(row=curr_row, column=1, value=no)
    c1.alignment = ALIGN_CENTER
    
    c2 = ws.cell(row=curr_row, column=2, value=name)
    c2.alignment = ALIGN_LEFT
    c2.font = FONT_BOLD
    
    c3 = ws.cell(row=curr_row, column=3, value=desc)
    c3.alignment = ALIGN_LEFT
    c3.font = FONT_REGULAR
    
    c4 = ws.cell(row=curr_row, column=4, value=qty)
    c4.alignment = ALIGN_CENTER
    
    c5 = ws.cell(row=curr_row, column=5, value=unit)
    c5.alignment = ALIGN_CENTER
    
    c6 = ws.cell(row=curr_row, column=6, value=price)
    c6.alignment = ALIGN_RIGHT
    c6.number_format = CURRENCY_FORMAT
    
    c7 = ws.cell(row=curr_row, column=7, value=f"=D{curr_row}*F{curr_row}")
    c7.alignment = ALIGN_RIGHT
    c7.font = FONT_BOLD
    c7.number_format = CURRENCY_FORMAT
    
    for c in range(1, 8):
        cell = ws.cell(row=curr_row, column=c)
        cell.border = BORDER_CELL
        if row_fill:
            cell.fill = row_fill

# --- 4. SUMMARY TOTALS ---
last_item_row = start_row + len(items) - 1

# Subtotal
r_subtotal = last_item_row + 1
ws.row_dimensions[r_subtotal].height = 24
ws.merge_cells(f"A{r_subtotal}:F{r_subtotal}")
cell_lbl1 = ws[f"A{r_subtotal}"]
cell_lbl1.value = "SUBTOTAL BIAYA FITUR"
cell_lbl1.alignment = ALIGN_RIGHT
cell_lbl1.font = FONT_BOLD

cell_val1 = ws.cell(row=r_subtotal, column=7, value=f"=SUM(G{start_row}:G{last_item_row})")
cell_val1.alignment = ALIGN_RIGHT
cell_val1.font = FONT_BOLD
cell_val1.number_format = CURRENCY_FORMAT

# Diskon Promo Fresh Graduate (0)
r_disc = r_subtotal + 1
ws.row_dimensions[r_disc].height = 22
ws.merge_cells(f"A{r_disc}:F{r_disc}")
cell_lbl2 = ws[f"A{r_disc}"]
cell_lbl2.value = "Diskon Special Promo Fresh Graduate SMK RPL"
cell_lbl2.alignment = ALIGN_RIGHT
cell_lbl2.font = FONT_REGULAR

cell_val2 = ws.cell(row=r_disc, column=7, value=0)
cell_val2.alignment = ALIGN_RIGHT
cell_val2.font = FONT_REGULAR
cell_val2.number_format = CURRENCY_FORMAT

# Grand Total
r_total = r_disc + 1
ws.row_dimensions[r_total].height = 28
ws.merge_cells(f"A{r_total}:F{r_total}")
cell_lbl3 = ws[f"A{r_total}"]
cell_lbl3.value = "GRAND TOTAL ESTIMASI BIAYA PROYEK"
cell_lbl3.alignment = ALIGN_RIGHT
cell_lbl3.font = FONT_TOTAL

cell_val3 = ws.cell(row=r_total, column=7, value=f"=G{r_subtotal}-G{r_disc}")
cell_val3.alignment = ALIGN_RIGHT
cell_val3.font = FONT_TOTAL
cell_val3.fill = COLOR_TOTAL_BG
cell_val3.number_format = CURRENCY_FORMAT

# Apply borders to summary section
for r in range(r_subtotal, r_total + 1):
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = BORDER_TOTAL

# --- 5. SKEMA PEMBAYARAN & TERMASUK KETENTUAN ---
r_terms = r_total + 2
ws.merge_cells(f"A{r_terms}:D{r_terms}")
t_title1 = ws[f"A{r_terms}"]
t_title1.value = "TERMIN PEMBAYARAN (PAYMENT MILESTONES)"
t_title1.font = FONT_SECTION

ws.merge_cells(f"E{r_terms}:G{r_terms}")
t_title2 = ws[f"E{r_terms}"]
t_title2.value = "LAYANAN TERMASUK (INCLUSIONS)"
t_title2.font = FONT_SECTION

terms_list = [
    ("Termin 1 (DP 30%)", "=G" + str(r_total) + "*0.3", "Source Code Lengkap (Laravel 13 & Tailwind)"),
    ("Termin 2 (Progress 40%)", "=G" + str(r_total) + "*0.4", "Dokumentasi Lengkap (Flowchart, ERD, DFD)"),
    ("Termin 3 (Pelunasan 30%)", "=G" + str(r_total) + "*0.3", "Free Maintenance & Bug Fix (1 Bulan)")
]

for idx, (t_lbl, t_formula, inc) in enumerate(terms_list, r_terms + 1):
    ws.row_dimensions[idx].height = 20
    
    ws.cell(row=idx, column=1, value=t_lbl).font = FONT_REGULAR
    ws.cell(row=idx, column=1).alignment = ALIGN_LEFT
    
    ws.cell(row=idx, column=2, value=t_formula).font = FONT_BOLD
    ws.cell(row=idx, column=2).alignment = ALIGN_RIGHT
    ws.cell(row=idx, column=2).number_format = CURRENCY_FORMAT
    
    ws.cell(row=idx, column=5, value="✓ " + inc).font = FONT_REGULAR
    ws.cell(row=idx, column=5).alignment = ALIGN_LEFT

# Auto-fit Column Widths
col_widths = {
    'A': 6,   # No
    'B': 38,  # Modul / Fitur
    'C': 65,  # Deskripsi Spesifikasi
    'D': 8,   # Qty
    'E': 10,  # Satuan
    'F': 20,  # Harga Satuan
    'G': 22   # Total Biaya
}

for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# Save Workbook
output_file = r"d:\laragon\www\salon-yalia-beauty\Rancangan_Harga_Website_Salon_Yalia_Beauty_SMK.xlsx"
wb.save(output_file)
print(f"Quotation Excel (SMK Rate) saved at: {output_file}")
